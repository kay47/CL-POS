from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, current_app, session
from flask_login import login_required, current_user
from app.models import Product, Sale, SaleItem, Expense
from app.decorators import manager_required, role_required
from app import db
from decimal import Decimal
from app.forms import SaleStatusForm
from datetime import datetime, date, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import func, and_, text

bp = Blueprint('pos', __name__)

@bp.route('/')
@login_required
def pos_page():
    products = Product.query.filter(Product.quantity > 0).order_by(Product.name).all()

    # Check if we're continuing a sale
    continue_sale_id = session.get('continue_sale_id')
    if continue_sale_id:
        session.pop('continue_sale_id', None)
        return render_template('pos/pos.html', products=products, continue_sale_id=continue_sale_id)
    
    return render_template('pos/pos.html', products=products)

@bp.route('/preview', methods=['POST'])
@login_required
def checkout_preview():
    """Show preview of items before completing the sale - includes retail units"""
    try:
        items = request.json.get('items', [])
        if not items:
            return jsonify({'error': 'No items in cart'}), 400
        
        preview_items = []
        total = Decimal('0.00')
        total_profit = Decimal('0.00')
        
        for item in items:
            product_id = int(item['product_id'])
            quantity = int(item['quantity'])
            unit_type = item.get('unit_type', 'full')
            
            product = Product.query.get(product_id)
            if not product:
                return jsonify({'error': f'Product not found'}), 400
            
            # Calculate inventory needed using the new method
            inventory_needed = product.calculate_inventory_deduction(quantity, unit_type)
            
            if product.quantity < inventory_needed:
                if unit_type == 'unit':
                    return jsonify({
                        'error': f'Insufficient stock for {product.name}. '
                                f'Need {quantity} units, have {int(product.total_units_available)} units available'
                    }), 400
                else:
                    return jsonify({
                        'error': f'Insufficient stock for {product.name}. Available: {product.quantity}'
                    }), 400
            
            price = product.get_price_for_unit(unit_type)
            line_total = Decimal(str(price)) * quantity
            total += line_total
            
            line_profit = Decimal(str(product.get_profit_for_unit(unit_type))) * quantity
            total_profit += line_profit
            
            preview_items.append({
                'product_id': product.id,
                'product_name': product.name,
                'sku': product.sku,
                'unit_type': unit_type,
                'price': float(price),
                'quantity': quantity,
                'line_total': float(line_total),
                'line_profit': float(line_profit),
                'available_stock': float(product.quantity),
                'available_units': int(product.total_units_available) if unit_type == 'unit' else None
            })
        
        return jsonify({
            'success': True,
            'items': preview_items,
            'total': float(total),
            'total_profit': float(total_profit),
            'item_count': len(preview_items)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@bp.route('/checkout', methods=['POST'])
@login_required
def checkout():
    """Complete the sale with fractional pricing and retail units support"""
    try:
        data = request.json
        items = data.get('items', [])
        sale_status = data.get('status', 'completed')
        payment_method = data.get('payment_method', 'cash')
        
        # Ensure amount_paid is properly converted to Decimal
        try:
            amount_paid = Decimal(str(data.get('amount_paid', 0)))
        except (ValueError, TypeError):
            amount_paid = Decimal('0')
            
        try:
            change_given = Decimal(str(data.get('change_given', 0)))
        except (ValueError, TypeError):
            change_given = Decimal('0')
        
        if not items:
            return jsonify({'error': 'No items in cart'}), 400
        
        if sale_status not in ['completed', 'pending']:
            sale_status = 'completed'
            
        if payment_method not in ['cash', 'card', 'mobile_money', 'bank_transfer']:
            payment_method = 'cash'

        continuing_sale_id = session.get('continue_sale_id')

        # STEP 1: Calculate total and validate stock FIRST (before any database changes)
        total = Decimal('0.00')
        total_profit = Decimal('0.00')
        validated_items = []
        
        for item in items:
            product_id = int(item['product_id'])
            quantity = int(item['quantity'])
            unit_type = item.get('unit_type', 'full')
            
            product = Product.query.get(product_id)
            if not product:
                raise ValueError(f'Product not found')
            
            # Calculate inventory deduction
            inventory_deduction = product.calculate_inventory_deduction(quantity, unit_type)
            
            # Check if there's enough inventory
            if product.quantity < inventory_deduction:
                if unit_type == 'unit':
                    raise ValueError(
                        f'Insufficient stock for {product.name}. '
                        f'Need {quantity} units ({float(inventory_deduction):.2f} packs), '
                        f'have {float(product.quantity * product.units_per_pack):.0f} units '
                        f'({float(product.quantity):.2f} packs)'
                    )
                else:
                    raise ValueError(
                        f'Insufficient stock for {product.name}. '
                        f'Need {inventory_deduction} packs, have {product.quantity}'
                    )
            
            # Get price and cost basis
            price = product.get_price_for_unit(unit_type)
            
            if unit_type == 'unit':
                cost_basis = product.purchase_price / Decimal(str(product.units_per_pack))
            elif unit_type == 'half':
                cost_basis = product.purchase_price / Decimal('2')
            elif unit_type == 'quarter':
                cost_basis = product.purchase_price / Decimal('4')
            else:  # full
                cost_basis = product.purchase_price
            
            line_total = Decimal(str(price)) * quantity
            line_profit = (Decimal(str(price)) - Decimal(str(cost_basis))) * quantity
            
            total += line_total
            total_profit += line_profit
            
            # Store validated item data for later processing
            validated_items.append({
                'product': product,
                'product_id': product_id,
                'quantity': quantity,
                'unit_type': unit_type,
                'price': price,
                'cost_basis': cost_basis,
                'inventory_deduction': inventory_deduction
            })
        
        # STEP 2: Validate payment amount for completed sales BEFORE any changes
        if sale_status == 'completed' and amount_paid < total:
            return jsonify({
                'error': f'Payment amount (GHS {float(amount_paid):.2f}) is less than total amount (GHS {float(total):.2f})'
            }), 400
        
        # STEP 3: Now proceed with sale creation/update
        if continuing_sale_id:
            sale = Sale.query.get(continuing_sale_id)
            if not sale or sale.status != 'pending':
                return jsonify({'error': 'Invalid pending sale'}), 400
            
            if current_user.role == 'cashier' and sale.clerk_id != current_user.id:
                return jsonify({'error': 'Access denied'}), 403
            
            # Restore inventory from previous sale items
            for item in sale.sale_items:
                inventory_to_restore = item.product.calculate_inventory_deduction(
                    item.quantity, item.unit_type
                )
                item.product.quantity += inventory_to_restore
                db.session.delete(item)
            
            sale.payment_method = payment_method
            sale.amount_paid = amount_paid
            sale.change_given = change_given
            sale.status = sale_status
            
            session.pop('continue_sale_id', None)
        else:
            sale = Sale(
                clerk_id=current_user.id, 
                total_amount=0,
                total_profit=Decimal('0.00'),
                status=sale_status,
                payment_method=payment_method,
                amount_paid=amount_paid,
                change_given=change_given
            )
            db.session.add(sale)
        
        db.session.flush()
        
        # STEP 4: Process validated items and update inventory
        for item_data in validated_items:
            product = item_data['product']
            
            # Update stock only if sale is completed
            if sale_status == 'completed':
                product.quantity -= item_data['inventory_deduction']
            
            # Create sale item
            sale_item = SaleItem(
                sale=sale,
                product=product,
                quantity=item_data['quantity'],
                unit_type=item_data['unit_type'],
                price_at_sale=item_data['price'],
                cost_at_sale=item_data['cost_basis']
            )
            db.session.add(sale_item)
        
        # Update sale totals
        sale.total_amount = float(total)
        sale.total_profit = float(total_profit)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'sale_id': sale.id,
            'invoice_number': sale.invoice_number,
            'total': float(total),
            'total_profit': float(total_profit),
            'payment_method': payment_method,
            'amount_paid': float(amount_paid),
            'change_given': float(change_given),
            'status': sale.status,
            'message': f'Sale {sale.invoice_number} completed successfully!'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@bp.route('/sales/<int:sale_id>/status', methods=['GET', 'POST'])
@login_required
@manager_required
def update_sale_status(sale_id):
    """Update sale status with CORRECT fractional inventory adjustments"""
    sale = Sale.query.get_or_404(sale_id)
    form = SaleStatusForm(obj=sale)
    
    if form.validate_on_submit():
        old_status = sale.status
        new_status = form.status.data
        
        # Handle inventory changes when status changes
        if old_status != new_status:
            if old_status == 'completed' and new_status in ['pending', 'cancelled']:
                # Return items to inventory
                for item in sale.sale_items:
                    inventory_to_restore = item.product.calculate_inventory_deduction(
                        item.quantity, item.unit_type
                    )
                    item.product.quantity += inventory_to_restore
            
            elif old_status in ['pending', 'cancelled'] and new_status == 'completed':
                # Remove items from inventory
                for item in sale.sale_items:
                    inventory_to_deduct = item.product.calculate_inventory_deduction(
                        item.quantity, item.unit_type
                    )
                    
                    if item.product.quantity < inventory_to_deduct:
                        if item.unit_type == 'unit':
                            flash(f'Insufficient stock for {item.product.name}. Need {item.quantity} units, have {int(item.product.total_units_available)} units. Cannot complete sale.', 'error')
                        else:
                            flash(f'Insufficient stock for {item.product.name}. Need {inventory_to_deduct} packs, have {item.product.quantity}. Cannot complete sale.', 'error')
                        return redirect(url_for('pos.update_sale_status', sale_id=sale_id))
                    
                    item.product.quantity -= inventory_to_deduct
        
        sale.status = new_status
        db.session.commit()
        
        flash(f'Sale #{sale.id} status updated from {old_status} to {new_status}', 'success')
        return redirect(url_for('pos.list_sales'))
    
    return render_template('pos/update_status.html', form=form, sale=sale)

@bp.route('/receipt/<int:sale_id>')
@login_required
def receipt(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    return render_template('pos/receipt.html', sale=sale)

@bp.route('/sales/report')
@login_required
@manager_required
def sales_report():
    """Generate sales report with profit analysis and expenses overlay"""
    # Get date range from query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
    if not end_date:
        end_date = datetime.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get sales within date range
    sales_query = Sale.query.filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    )
    
    # Calculate totals
    total_sales = sales_query.count()
    total_revenue = sales_query.with_entities(func.sum(Sale.total_amount)).scalar() or 0
    total_profit = sales_query.with_entities(func.sum(Sale.total_profit)).scalar() or 0
    
    # Get daily revenue data
    daily_revenue = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.sum(Sale.total_amount).label('revenue'),
        func.sum(Sale.total_profit).label('profit')
    ).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(func.date(Sale.created_at)).order_by('date').all()
    
    # Get daily expenses data
    daily_expenses = db.session.query(
        Expense.date.label('date'),
        func.sum(Expense.amount).label('total_expense')
    ).filter(
        and_(
            Expense.date >= start_date,
            Expense.date <= end_date
        )
    ).group_by(Expense.date).all()
    
    # Create expense lookup dictionary
    expense_dict = {exp.date: float(exp.total_expense) for exp in daily_expenses}
    
    # Combine revenue and expense data
    daily_data = []
    for row in daily_revenue:
        date_obj = row.date if isinstance(row.date, date) else datetime.strptime(row.date, '%Y-%m-%d').date()
        daily_data.append({
            'date': date_obj.strftime('%Y-%m-%d'),
            'revenue': float(row.revenue),
            'profit': float(row.profit),
            'expense': expense_dict.get(date_obj, 0.0)
        })
    
    # Calculate total expenses in period
    total_expenses = sum(expense_dict.values())
    
    # Calculate net profit (profit - expenses)
    net_profit = float(total_profit) - total_expenses
    
    # Get profit breakdown by unit type
    unit_type_stats = db.session.query(
        SaleItem.unit_type,
        func.count(SaleItem.id).label('count'),
        func.sum(SaleItem.quantity).label('total_quantity'),
        func.sum(SaleItem.price_at_sale * SaleItem.quantity).label('total_revenue'),
        func.sum((SaleItem.price_at_sale - SaleItem.cost_at_sale) * SaleItem.quantity).label('total_profit')
    ).join(Sale).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(SaleItem.unit_type).all()
    
    # Get payment method breakdown
    payment_stats = db.session.query(
        Sale.payment_method,
        func.count(Sale.id).label('count'),
        func.sum(Sale.total_amount).label('total_amount')
    ).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(Sale.payment_method).all()
    
    return render_template('pos/sales_report.html', 
                         start_date=start_date, 
                         end_date=end_date,
                         total_sales=total_sales,
                         total_revenue=total_revenue,
                         total_profit=total_profit,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         daily_data=daily_data,
                         unit_type_stats=unit_type_stats,
                         payment_stats=payment_stats)

@bp.route('/export-sales-excel')
@login_required
@role_required(['clerk', 'manager', 'admin'])
def export_sales_excel():
    """Export sales data to Excel format"""
    try:
        status_filter = request.args.get('status', 'all')
        
        query = Sale.query
        if status_filter != 'all':
            query = query.filter(Sale.status == status_filter)
        
        sales = query.order_by(Sale.created_at.desc()).all()
        
        if not sales:
            flash('No sales data to export', 'warning')
            return redirect(url_for('pos.list_sales'))
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Sales Report - {status_filter.title()}"
        
        headers = [
            'Date', 'Time', 'Invoice Number', 'Items Count', 'Total Amount', 
            'Amount Paid', 'Change Given', 'Payment Method', 'Status', 'Clerk',
            'Product Name', 'SKU', 'Unit Type', 'Quantity', 'Unit Price', 'Line Total'
        ]
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row = 2
        for sale in sales:
            if sale.sale_items:
                for i, item in enumerate(sale.sale_items):
                    if i == 0:
                        ws.cell(row=row, column=1, value=sale.created_at.strftime('%Y-%m-%d'))
                        ws.cell(row=row, column=2, value=sale.created_at.strftime('%H:%M:%S'))
                        ws.cell(row=row, column=3, value=sale.invoice_number)
                        ws.cell(row=row, column=4, value=len(sale.sale_items))
                        ws.cell(row=row, column=5, value=float(sale.total_amount))
                        ws.cell(row=row, column=6, value=float(sale.amount_paid))
                        ws.cell(row=row, column=7, value=float(sale.change_given))
                        ws.cell(row=row, column=8, value=sale.payment_method.title())
                        ws.cell(row=row, column=9, value=sale.status.title())
                        ws.cell(row=row, column=10, value=sale.clerk.username)
                    
                    ws.cell(row=row, column=11, value=item.product.name)
                    ws.cell(row=row, column=12, value=item.product.sku)
                    ws.cell(row=row, column=13, value=item.unit_display)
                    ws.cell(row=row, column=14, value=item.quantity)
                    ws.cell(row=row, column=15, value=float(item.price_at_sale))
                    ws.cell(row=row, column=16, value=float(item.line_total))
                    
                    row += 1
            else:
                ws.cell(row=row, column=1, value=sale.created_at.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=sale.created_at.strftime('%H:%M:%S'))
                ws.cell(row=row, column=3, value=sale.invoice_number)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=float(sale.total_amount))
                ws.cell(row=row, column=6, value=float(sale.amount_paid))
                ws.cell(row=row, column=7, value=float(sale.change_given))
                ws.cell(row=row, column=8, value=sale.payment_method.title())
                ws.cell(row=row, column=9, value=sale.status.title())
                ws.cell(row=row, column=10, value=sale.clerk.username)
                row += 1
        
        summary_row = row + 1
        ws.cell(row=summary_row, column=1, value="TOTALS:")
        ws.cell(row=summary_row, column=4, value=sum(len(sale.sale_items) for sale in sales))
        ws.cell(row=summary_row, column=5, value=sum(float(sale.total_amount) for sale in sales))
        ws.cell(row=summary_row, column=6, value=sum(float(sale.amount_paid) for sale in sales))
        ws.cell(row=summary_row, column=7, value=sum(float(sale.change_given) for sale in sales))
        
        for col in range(1, 17):
            cell = ws.cell(row=summary_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"sales_report_{status_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Export error: {str(e)}")
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('pos.list_sales'))

@bp.route('/continue-sale/<int:sale_id>')
@login_required
def continue_sale(sale_id):
    """Continue shopping on a pending sale"""
    try:
        sale = Sale.query.get_or_404(sale_id)
        
        if sale.status != 'pending':
            flash('Can only continue pending sales', 'warning')
            return redirect(url_for('pos.list_sales'))
        
        if current_user.role == 'clerk' and sale.clerk_id != current_user.id:
            flash('You can only continue your own sales', 'danger')
            return redirect(url_for('pos.list_sales'))
        
        session['continue_sale_id'] = sale_id
        flash(f'Continuing sale {sale.invoice_number}', 'success')
        
        return redirect(url_for('pos.pos_page'))
        
    except Exception as e:
        current_app.logger.error(f"Continue sale error: {str(e)}")
        flash('Failed to continue sale', 'danger')
        return redirect(url_for('pos.list_sales'))

@bp.route('/load-pending-sale/<int:sale_id>')
@login_required
def load_pending_sale(sale_id):
    """API endpoint to load pending sale items into cart"""
    try:
        sale = Sale.query.get_or_404(sale_id)
        
        if sale.status != 'pending':
            return jsonify({'success': False, 'error': 'Sale is not pending'})
        
        if current_user.role == 'clerk' and sale.clerk_id != current_user.id:
            return jsonify({'success': False, 'error': 'Access denied'})
        
        cart_items = []
        for item in sale.sale_items:
            # Calculate available stock including reserved quantity
            if item.unit_type == 'unit':
                reserved_packs = item.product.calculate_inventory_deduction(item.quantity, 'unit')
                available_units = int((item.product.quantity + reserved_packs) * item.product.units_per_pack)
                stock_value = item.product.quantity + reserved_packs
            else:
                reserved_packs = item.product.calculate_inventory_deduction(item.quantity, item.unit_type)
                stock_value = item.product.quantity + reserved_packs
                available_units = None
            
            cart_items.append({
                'id': item.product.id,
                'name': item.product.name,
                'price': float(item.price_at_sale),
                'quantity': item.quantity,
                'stock': float(stock_value),
                'sku': item.product.sku,
                'unit_type': item.unit_type,
                'profit': float(item.line_profit / item.quantity) if item.quantity > 0 else 0,
                'units_per_pack': item.product.units_per_pack,
                'total_units': available_units
            })
        
        return jsonify({
            'success': True,
            'cart_items': cart_items,
            'sale_info': {
                'id': sale.id,
                'invoice_number': sale.invoice_number,
                'payment_method': sale.payment_method,
                'amount_paid': float(sale.amount_paid)
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Load pending sale error: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to load sale'})

@bp.route('/list-sales')
@login_required
def list_sales():
    """Display list of sales with filtering and pagination"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Sale.query
    
    if current_user.role == 'clerk':
        query = query.filter(Sale.clerk_id == current_user.id)
    
    if status_filter != 'all':
        query = query.filter(Sale.status == status_filter)
    
    sales = query.order_by(Sale.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template(
        'pos/sales_list.html', 
        sales=sales, 
        status_filter=status_filter,
        current_time=datetime.now()
    )

@bp.route('/profits-dashboard')
@login_required
@manager_required
def profits_dashboard():
    """Comprehensive profits dashboard"""
    from datetime import timedelta
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    period = request.args.get('period', '30')
    
    if not start_date:
        start_date = (datetime.now() - timedelta(days=int(period))).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
    if not end_date:
        end_date = datetime.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    total_sales_query = Sale.query.filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    )
    
    total_revenue = total_sales_query.with_entities(func.sum(Sale.total_amount)).scalar() or 0
    total_profit = total_sales_query.with_entities(func.sum(Sale.total_profit)).scalar() or 0
    total_sales_count = total_sales_query.count()
    
    total_cost = float(total_revenue) - float(total_profit)
    profit_margin = (float(total_profit) / float(total_revenue) * 100) if total_revenue > 0 else 0
    
    daily_profits_raw = db.session.query(
        func.date(Sale.created_at).label('date'),
        func.sum(Sale.total_amount).label('revenue'),
        func.sum(Sale.total_profit).label('profit'),
        func.count(Sale.id).label('sales_count')
    ).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(func.date(Sale.created_at)).order_by('date').all()
    
    daily_profits = []
    for row in daily_profits_raw:
        if isinstance(row.date, str):
            date_obj = datetime.strptime(row.date, '%Y-%m-%d').date()
        else:
            date_obj = row.date
            
        daily_profits.append({
            'date': date_obj,
            'revenue': row.revenue,
            'profit': row.profit,
            'sales_count': row.sales_count
        })
    
    product_profits = db.session.query(
        Product.name.label('product_name'),
        Product.sku.label('sku'),
        func.sum(SaleItem.quantity).label('quantity_sold'),
        func.sum((SaleItem.price_at_sale - SaleItem.cost_at_sale) * SaleItem.quantity).label('total_profit'),
        func.sum(SaleItem.price_at_sale * SaleItem.quantity).label('total_revenue')
    ).join(Sale).join(Product).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(Product.id, Product.name, Product.sku).order_by(text('total_profit DESC')).limit(10).all()
    
    unit_profits = db.session.query(
        SaleItem.unit_type,
        func.sum(SaleItem.quantity).label('quantity_sold'),
        func.sum((SaleItem.price_at_sale - SaleItem.cost_at_sale) * SaleItem.quantity).label('total_profit'),
        func.sum(SaleItem.price_at_sale * SaleItem.quantity).label('total_revenue')
    ).join(Sale).filter(
        and_(
            Sale.created_at >= start_date,
            Sale.created_at <= end_date,
            Sale.status == 'completed'
        )
    ).group_by(SaleItem.unit_type).all()
    
    monthly_profits = []
    if (end_date - start_date).days > 30:
        monthly_profits_raw = db.session.query(
            func.strftime('%Y-%m', Sale.created_at).label('month'),
            func.sum(Sale.total_amount).label('revenue'),
            func.sum(Sale.total_profit).label('profit'),
            func.count(Sale.id).label('sales_count')
        ).filter(
            and_(
                Sale.created_at >= start_date,
                Sale.created_at <= end_date,
                Sale.status == 'completed'
            )
        ).group_by(func.strftime('%Y-%m', Sale.created_at)).order_by('month').all()
        
        monthly_profits = []
        for row in monthly_profits_raw:
            monthly_profits.append({
                'month': row.month,
                'revenue': row.revenue,
                'profit': row.profit,
                'sales_count': row.sales_count
            })
    
    return render_template('pos/profits_dashboard.html',
                         start_date=start_date,
                         end_date=end_date,
                         total_revenue=total_revenue,
                         total_profit=total_profit,
                         total_cost=total_cost,
                         total_sales_count=total_sales_count,
                         profit_margin=profit_margin,
                         daily_profits=daily_profits,
                         product_profits=product_profits,
                         unit_profits=unit_profits,
                         monthly_profits=monthly_profits,
                         period=period)